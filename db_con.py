import sqlite3
from datetime import datetime
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os

def connect_to_database():
    db_path = os.path.join(os.path.dirname(__file__), 'visitor_attendance.db')  # Path to your SQLite database file
    return sqlite3.connect(db_path)

# Registration of Account Page___________________________________________________________________________________________
def register_security_admin(name, username, password, register_window, FnExistlabel, UnExistlabel):
    conn = connect_to_database()
    cursor = conn.cursor()
    try:
        # Check if the full name already exists in the database
        cursor.execute("SELECT COUNT(*) FROM security_admin WHERE LOWER(sec_name) = LOWER(?)", (name,))
        if cursor.fetchone()[0] > 0:
            FnExistlabel.configure(text='Name is already taken.')
            return False  # Early exit if the name exists

        # Check if the username already exists in the database
        cursor.execute("SELECT COUNT(*) FROM security_admin WHERE LOWER(sec_username) = LOWER(?)", (username,))
        if cursor.fetchone()[0] > 0:
            UnExistlabel.configure(text='Username is already taken.')
            return False  # Early exit if the username exists

        # If neither name nor username is taken, proceed to insert
        query_insert = "INSERT INTO security_admin (sec_name, sec_username, sec_password) VALUES (?, ?, ?)"
        cursor.execute(query_insert, (name, username, password))
        conn.commit()
        return True
    except sqlite3.Error as err:
        print(f"Failed to insert into security_admin: {err}")
        return False
    finally:
        cursor.close()
        conn.close()

# Sign In Page___________________________________________________________________________________________
def validate_login_credentials(username, password):
    conn = connect_to_database()
    cursor = conn.cursor()

    query = "SELECT sec_id FROM security_admin WHERE sec_username = ? AND sec_password = ?"
    cursor.execute(query, (username, password))

    # Fetch the result of the query
    result = cursor.fetchone()

    # Close cursor and connection
    cursor.close()
    conn.close()

    # Check if result is not None
    if result is not None:
        # Extract sec_id from the result tuple
        sec_id = result[0]
        # Return a boolean indicating success and sec_id
        return True, sec_id
    else:
        # Handle the case when the query returns None (no result found)
        return False, None

# Set New Password___________________________________________________________________________________________
def validate_and_update_password(username, new_password):
    conn = connect_to_database()
    cursor = conn.cursor()

    # Validate username
    query_validate = "SELECT sec_id FROM security_admin WHERE sec_username = ?"
    cursor.execute(query_validate, (username,))
    result = cursor.fetchone()

    if result is None:
        cursor.close()
        conn.close()
        return False, "Invalid Username"

    # Update password
    query_update = "UPDATE security_admin SET sec_password = ? WHERE sec_username = ?"
    cursor.execute(query_update, (new_password, username))
    conn.commit()

    updated = cursor.rowcount > 0

    cursor.close()
    conn.close()

    if updated:
        return True, "Password Updated Successfully"
    else:
        return False, "Password update failed"

# Home Page___________________________________________________________________________________________
def count_logged_in():
    conn = connect_to_database()
    cursor = conn.cursor()
    try:
        today = datetime.now().date()
        cursor.execute("SELECT COUNT(*) FROM visitor_data WHERE log_stat = 1 AND log_day = ?", (today,))
        count = cursor.fetchone()[0]
        return count
    except sqlite3.Error as err:
        print(f"Failed to count logged in visitors: {err}")
        return 0
    finally:
        cursor.close()
        conn.close()

def count_logged_out():
    conn = connect_to_database()
    cursor = conn.cursor()
    try:
        today = datetime.now().date()
        cursor.execute("SELECT COUNT(*) FROM visitor_data WHERE log_stat = 0 AND log_day = ?", (today,))
        count = cursor.fetchone()[0]
        return count
    except sqlite3.Error as err:
        print(f"Failed to count logged out visitors: {err}")
        return 0
    finally:
        cursor.close()
        conn.close()

def count_total_today():
    conn = connect_to_database()
    cursor = conn.cursor()
    try:
        today = datetime.now().date()
        cursor.execute("SELECT COUNT(*) FROM visitor_data WHERE log_day = ?", (today,))
        count = cursor.fetchone()[0]
        return count
    except sqlite3.Error as err:
        print(f"Failed to count total visitors today: {err}")
        return 0
    finally:
        cursor.close()
        conn.close()

# Log in Visitor Page___________________________________________________________________________________________
def insert_visitor_data(visit_name, res_id, log_purpose, sec_id):
    conn = connect_to_database()
    cursor = conn.cursor()
    try:
        # Prepare the current date and time
        current_datetime = datetime.now()
        log_day = current_datetime.date()
        login_time = current_datetime.strftime('%H:%M:%S')

        # Check if the visitor with the same name is already logged in
        query_check_existing = """
        SELECT * FROM visitor_data 
        WHERE visit_name = ? AND sec_id = ? AND log_stat = TRUE AND logout_time IS NULL
        """
        cursor.execute(query_check_existing, (visit_name, sec_id))
        existing_visitor = cursor.fetchone()

        if existing_visitor:
            # Visitor is already logged in
            success = False
        else:
            # Prepare SQL query to insert data
            query_insert = """
            INSERT INTO visitor_data (
                visit_name,
                res_id,
                log_purpose,
                log_day,
                login_time,
                log_stat,
                sec_id
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """
            # Values to be inserted
            data_tuple = (visit_name, res_id, log_purpose, log_day, login_time, True, sec_id)

            # Execute the query
            cursor.execute(query_insert, data_tuple)

            # Commit changes
            conn.commit()
            success = True
    except sqlite3.Error as err:
        print(f"Failed to insert visitor data: {err}")
        success = False
    finally:
        cursor.close()
        conn.close()
    return success

def fetch_residents():
    conn = connect_to_database()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT res_id, res_address FROM resident_data")
        residents = cursor.fetchall()  # Fetch all rows
        return residents
    except sqlite3.Error as err:
        print(f"Failed to fetch resident data: {err}")
        return []
    finally:
        cursor.close()
        conn.close()

# Log out Visitor Page___________________________________________________________________________________________
def logout_visitor(visit_name, sec_id, Existinglabel, logoutbtn):
    conn = connect_to_database()
    if conn is None:
        Existinglabel.configure(text='Failed to connect to the database.')
        return

    cursor = conn.cursor()
    logoutbtn.configure(state="disabled")
    try:
        query_check = """
        SELECT login_time, logout_time, log_day, log_stat 
        FROM visitor_data 
        WHERE visit_name = ? AND sec_id = ? 
        ORDER BY log_day DESC, login_time DESC 
        LIMIT 1
        """
        cursor.execute(query_check, (visit_name, sec_id))
        result = cursor.fetchone()

        if result:
            login_time, logout_time, log_day, log_stat = result
            current_datetime = datetime.now()
            logout_time_new = current_datetime.strftime('%H:%M:%S')
            log_day_new = current_datetime.date()

            if login_time is not None and logout_time is None and log_stat:
                query_update = """
                UPDATE visitor_data 
                SET logout_time = ?, log_day = ?, log_stat = FALSE 
                WHERE visit_name = ? AND login_time = ? AND sec_id = ? AND log_stat = TRUE
                """
                cursor.execute(query_update, (logout_time_new, log_day_new, visit_name, login_time, sec_id))
                conn.commit()

                query_fetch = """
                SELECT vd.visit_name, vd.log_day, vd.login_time, vd.logout_time, rd.res_address, sa.sec_name, vd.log_purpose
                FROM visitor_data vd
                JOIN resident_data rd ON vd.res_id = rd.res_id
                JOIN security_admin sa ON vd.sec_id = sa.sec_id
                WHERE vd.visit_name = ? AND vd.sec_id = ? AND vd.login_time = ?
                """
                cursor.execute(query_fetch, (visit_name, sec_id, login_time))
                visitor_data = cursor.fetchone()

                if visitor_data:
                    save_data_to_excel(visitor_data)
                return True  # Logout was successful
            else:
                Existinglabel.configure(text='Log in First!')
        else:
            Existinglabel.configure(text='Log in first!')

    except Exception as e:
        print(f"An error occurred: {e}")
        Existinglabel.configure(text='An error occurred while processing your request.')
    finally:
        cursor.close()
        conn.close()

def save_data_to_excel(data):
    COL_NAMES = ['VISITOR NAME', 'DATE', 'LOGIN TIME', 'LOGOUT TIME', 'RESIDENT', 'SECURITY', 'PURPOSE']
    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    folder_path = os.path.join(desktop_path, 'Visitor_Attendance')
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    file_path = os.path.join(folder_path, f"{data[1]}_VAttendance.xlsx")
    
    # Check if file exists
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(COL_NAMES)  # Add header if the file is newly created

    # Add data row
    formatted_data = [item for item in data]
    sheet.append(formatted_data)

    # Adjust column widths
    for col_num, col_name in enumerate(COL_NAMES, 1):
        column_letter = get_column_letter(col_num)
        max_length = max(len(str(item)) for item in [col_name] + [formatted_data[col_num-1]]) + 2
        sheet.column_dimensions[column_letter].width = max_length

    workbook.save(file_path)

# Visitor Page___________________________________________________________________________________________
def fetch_visitor_data_desc(offset=0):
    conn = connect_to_database()
    cursor = conn.cursor()
    try:
        query = """
        SELECT vd.visit_name, vd.log_day, vd.login_time, vd.logout_time, rd.res_address, sa.sec_name, vd.log_purpose
        FROM visitor_data vd
        JOIN resident_data rd ON vd.res_id = rd.res_id
        JOIN security_admin sa ON vd.sec_id = sa.sec_id
        ORDER BY vd.log_day DESC, vd.login_time DESC
        LIMIT 15 OFFSET ?
        """
        cursor.execute(query, (offset,))
        data = cursor.fetchall()
        return data
    except sqlite3.Error as err:
        print(f"Failed to fetch visitor data: {err}")
        return []
    finally:
        cursor.close()
        conn.close()

def fetch_visitor_data_asc(offset=0):
    conn = connect_to_database()
    cursor = conn.cursor()
    try:
        query = """
        SELECT vd.visit_name, vd.log_day, vd.login_time, vd.logout_time, rd.res_address, sa.sec_name, vd.log_purpose
        FROM visitor_data vd
        JOIN resident_data rd ON vd.res_id = rd.res_id
        JOIN security_admin sa ON vd.sec_id = sa.sec_id
        ORDER BY vd.log_day ASC, vd.login_time ASC
        LIMIT 15 OFFSET ?
        """
        cursor.execute(query, (offset,))
        data = cursor.fetchall()
        return data
    except sqlite3.Error as err:
        print(f"Failed to fetch visitor data: {err}")
        return []
    finally:
        cursor.close()
        conn.close()

def fetch_visitor_data_name_asc(offset=0):
    conn = connect_to_database()
    cursor = conn.cursor()
    try:
        query = """
        SELECT vd.visit_name, vd.log_day, vd.login_time, vd.logout_time, rd.res_address, sa.sec_name, vd.log_purpose
        FROM visitor_data vd
        JOIN resident_data rd ON vd.res_id = rd.res_id
        JOIN security_admin sa ON vd.sec_id = sa.sec_id
        ORDER BY vd.visit_name ASC
        LIMIT 15 OFFSET ?
        """
        cursor.execute(query, (offset,))
        data = cursor.fetchall()
        return data
    except sqlite3.Error as err:
        print(f"Failed to fetch visitor data: {err}")
        return []
    finally:
        cursor.close()
        conn.close()

def fetch_visitor_data_name_desc(offset=0):
    conn = connect_to_database()
    cursor = conn.cursor()
    try:
        query = """
        SELECT vd.visit_name, vd.log_day, vd.login_time, vd.logout_time, rd.res_address, sa.sec_name, vd.log_purpose
        FROM visitor_data vd
        JOIN resident_data rd ON vd.res_id = rd.res_id
        JOIN security_admin sa ON vd.sec_id = sa.sec_id
        ORDER BY vd.visit_name DESC
        LIMIT 15 OFFSET ?
        """
        cursor.execute(query, (offset,))
        data = cursor.fetchall()
        return data
    except sqlite3.Error as err:
        print(f"Failed to fetch visitor data: {err}")
        return []
    finally:
        cursor.close()
        conn.close()

def get_total_visitors():
    conn = connect_to_database()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) FROM visitor_data")
        result = cursor.fetchone()
        total_visitors = result[0] if result else 0
        return total_visitors
    except sqlite3.Error as err:
        print(f"Failed to count total visitors: {err}")
        return 0
    finally:
        cursor.close()
        conn.close()

# Resident Page___________________________________________________________________________________________
def fetch_resident_data(offset=0, search_query=""):
    conn = connect_to_database()
    cursor = conn.cursor()
    try:
        if search_query:
            query = """
            SELECT res_id, res_name, res_address, res_phonenumber 
            FROM resident_data 
            WHERE res_name LIKE ? OR res_address LIKE ? OR res_phonenumber LIKE ? 
            LIMIT 15 OFFSET ?
            """
            cursor.execute(query, ('%' + search_query + '%', '%' + search_query + '%', '%' + search_query + '%', offset))
        else:
            query = "SELECT res_id, res_name, res_address, res_phonenumber FROM resident_data LIMIT 15 OFFSET ?"
            cursor.execute(query, (offset,))

        data = cursor.fetchall()

        cursor.execute("SELECT COUNT(*) FROM resident_data")
        total_results = cursor.fetchone()[0]

        return data, total_results
    except sqlite3.Error as err:
        print(f"Failed to fetch resident data: {err}")
        return [], 0
    finally:
        cursor.close()
        conn.close()

def get_total_residents():
    conn = connect_to_database()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) FROM resident_data")
        result = cursor.fetchone()
        total_residents = result[0] if result else 0
        return total_residents
    except sqlite3.Error as err:
        print(f"Failed to count total residents: {err}")
        return 0
    finally:
        cursor.close()
        conn.close()

def update_resident_data(window, res_id, name, address, phone):
    try:
        conn = connect_to_database()  # Ensure this function returns a valid connection
        cursor = conn.cursor()
        query = "UPDATE resident_data SET res_name = ?, res_address = ?, res_phonenumber = ? WHERE res_id = ?"
        cursor.execute(query, (name, address, phone, res_id))
        conn.commit()
    except Exception as e:  # Broad exception handling for any database-related errors
        print(f"Failed to update resident data: {e}")
        if conn:
            conn.rollback()
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
